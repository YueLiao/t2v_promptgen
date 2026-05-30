"""Multi-format prompt-list loader.

Supports: JSON / JSONL / TXT / CSV / XLSX.

Public API:
  - detect_format(filename, head_bytes) -> 'json'|'jsonl'|'txt'|'csv'|'xlsx'
  - load_prompts(file_bytes, filename, sheet_name=None) -> (SourceFile, rows)

All failures raise ParseError with a code from §11 of the design doc.
"""
from __future__ import annotations

import csv
import io
import json
from datetime import datetime
from typing import Literal

from ..core.rewrite_schema import (
    MAX_FILE_BYTES,
    MAX_ROWS,
    ParseError,
    SourceFile,
    SourceFormat,
)
from .encoding_detect import detect_and_decode


# ---------------------------------------------------------------------------
# Format detection
# ---------------------------------------------------------------------------

# Magic bytes / signatures for binary formats
_XLSX_SIG = b"PK\x03\x04"     # xlsx is a zip
_ZIP_SIG = b"PK"               # generic zip

# Map of suffix → format
_SUFFIX_MAP: dict[str, SourceFormat] = {
    "json": "json",
    "jsonl": "jsonl",
    "ndjson": "jsonl",
    "txt": "txt",
    "csv": "csv",
    "tsv": "csv",
    "xlsx": "xlsx",
}


def detect_format(filename: str, head_bytes: bytes) -> SourceFormat:
    """Identify file format from suffix and head bytes.

    Priority:
      1. Binary magic bytes win unconditionally (xlsx is always xlsx)
      2. Suffix decides between text formats
      3. Heuristic on head_bytes if suffix unknown
    """
    head = head_bytes[:512].lstrip()

    # Magic bytes always win for binary formats
    if head_bytes.startswith(_XLSX_SIG):
        return "xlsx"

    suffix = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

    # Suffix says binary but head is not — mismatch
    if suffix == "xlsx" and not head_bytes.startswith(_ZIP_SIG):
        raise ParseError("PARSE_FORMAT_MISMATCH", "后缀 .xlsx 但文件不是 Excel 格式")

    if suffix in _SUFFIX_MAP:
        return _SUFFIX_MAP[suffix]

    # Suffix unknown — heuristic on head
    if head.startswith(b"{") or head.startswith(b"["):
        # Could be JSON or JSONL — count newlines in a small window
        sample = head_bytes[:4096]
        # JSONL usually has multiple lines each starting with {
        lines = [l for l in sample.split(b"\n") if l.strip()]
        if len(lines) >= 2 and all(l.lstrip().startswith(b"{") for l in lines[:3]):
            return "jsonl"
        return "json"
    if b"," in head[:200] and b"\n" in head[:512]:
        return "csv"
    if head:
        return "txt"
    raise ParseError("PARSE_FORMAT_UNKNOWN", "文件格式无法识别")


# ---------------------------------------------------------------------------
# Per-format parsers — each returns list[dict]
# ---------------------------------------------------------------------------

def _parse_json(text: str) -> list[dict]:
    """Parse JSON. Accept array root or {prompts: [...]} dict root."""
    try:
        obj = json.loads(text)
    except json.JSONDecodeError as exc:
        raise ParseError(
            "PARSE_JSON_INVALID",
            f"JSON 解析失败: {exc.msg}",
            location=f"行 {exc.lineno} 列 {exc.colno}",
        ) from exc

    if isinstance(obj, list):
        rows = obj
    elif isinstance(obj, dict):
        # Common patterns: {"prompts": [...]} or {"data": [...]}
        for key in ("prompts", "data", "items", "rows"):
            if key in obj and isinstance(obj[key], list):
                rows = obj[key]
                break
        else:
            raise ParseError(
                "EMPTY_FILE",
                "JSON 根节点是 dict 且没有 prompts/data/items/rows 数组键",
            )
    else:
        raise ParseError("EMPTY_FILE", f"JSON 根节点类型 {type(obj).__name__} 无法处理")

    # Normalize each entry to dict; strings become {"text": "..."}
    out: list[dict] = []
    for i, item in enumerate(rows):
        if isinstance(item, dict):
            out.append(item)
        elif isinstance(item, str):
            out.append({"text": item})
        else:
            # skip non-dict, non-string entries with a warning
            out.append({"_invalid_row": str(type(item).__name__), "_row_idx": i})
    return out


def _parse_jsonl(text: str) -> list[dict]:
    out: list[dict] = []
    invalid_lines: list[int] = []
    for i, raw in enumerate(text.splitlines(), start=1):
        line = raw.strip()
        if not line:
            continue
        try:
            obj = json.loads(line)
        except json.JSONDecodeError:
            invalid_lines.append(i)
            continue
        if isinstance(obj, dict):
            out.append(obj)
        elif isinstance(obj, str):
            out.append({"text": obj})
        else:
            invalid_lines.append(i)
    if invalid_lines and not out:
        raise ParseError(
            "PARSE_JSON_INVALID",
            f"JSONL 全部 {len(invalid_lines)} 行解析失败",
        )
    if invalid_lines:
        # Soft warning attached to the last good row's metadata won't surface to UI;
        # the UI will see row_count smaller than line count.
        pass
    return out


def _parse_txt(text: str) -> list[dict]:
    """Plain text: each non-empty line is one prompt.

    Blank-line-separated paragraphs merge into single rows (lets users
    upload paragraph-style prompts).
    """
    lines = text.splitlines()
    if not lines:
        raise ParseError("EMPTY_FILE", "文本文件为空")

    # If document has blank-line separators, treat as paragraph mode
    has_blank_separator = any(not l.strip() for l in lines)
    if has_blank_separator:
        # Paragraph mode
        out: list[dict] = []
        buf: list[str] = []
        for line in lines:
            if line.strip():
                buf.append(line.strip())
            else:
                if buf:
                    out.append({"text": " ".join(buf)})
                    buf = []
        if buf:
            out.append({"text": " ".join(buf)})
    else:
        # Line mode
        out = [{"text": l.strip()} for l in lines if l.strip()]

    if not out:
        raise ParseError("EMPTY_FILE", "文本文件没有非空行")
    return out


def _parse_csv(text: str) -> list[dict]:
    """CSV / TSV. Auto-detect delimiter."""
    # Strip UTF-8 BOM if present in first char
    if text.startswith("﻿"):
        text = text[1:]

    # Sniff delimiter from first 4KB
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t;|")
    except csv.Error:
        # Default to comma if sniffing fails
        dialect = csv.excel

    try:
        reader = csv.DictReader(io.StringIO(text), dialect=dialect)
        out = list(reader)
    except csv.Error as exc:
        raise ParseError("PARSE_JSON_INVALID", f"CSV 解析失败: {exc}") from exc

    if not out:
        raise ParseError("EMPTY_FILE", "CSV 只有表头,没有数据行")
    return out


def _parse_xlsx(data: bytes, sheet_name: str | None) -> tuple[list[dict], str]:
    """Excel. Returns (rows, actual_sheet_name)."""
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError as exc:
        raise ParseError(
            "INTERNAL",
            "服务器缺少 openpyxl,无法解析 Excel。pip install openpyxl",
        ) from exc

    try:
        wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    except Exception as exc:
        raise ParseError(
            "PARSE_FORMAT_MISMATCH",
            f"Excel 文件损坏: {type(exc).__name__}",
        ) from exc

    sheets = wb.sheetnames
    if not sheets:
        raise ParseError("PARSE_XLSX_NO_SHEET", "Excel 文件没有任何 sheet")

    selected = sheet_name if (sheet_name and sheet_name in sheets) else sheets[0]
    ws = wb[selected]

    # Find header (first row with at least 1 non-empty cell)
    rows_iter = ws.iter_rows(values_only=True)
    header: list[str] = []
    for row in rows_iter:
        if any(cell is not None and str(cell).strip() for cell in row):
            header = [
                (str(c).strip() if c is not None else f"col_{i+1}")
                for i, c in enumerate(row)
            ]
            break
    if not header:
        raise ParseError("EMPTY_FILE", "Excel sheet 完全为空")

    # De-duplicate header names
    seen: dict[str, int] = {}
    for i, h in enumerate(header):
        if h in seen:
            seen[h] += 1
            header[i] = f"{h}_{seen[h]}"
        else:
            seen[h] = 0

    out: list[dict] = []
    for row in rows_iter:
        if all(cell is None or str(cell).strip() == "" for cell in row):
            continue   # skip empty rows
        record = {}
        for i, cell in enumerate(row):
            col_name = header[i] if i < len(header) else f"col_{i+1}"
            record[col_name] = cell if cell is not None else None
        out.append(record)
    if not out:
        raise ParseError("EMPTY_FILE", "Excel 只有表头,没有数据行")
    return out, selected


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def load_prompts(
    file_bytes: bytes,
    filename: str,
    sheet_name: str | None = None,
) -> tuple[SourceFile, list[dict]]:
    """Parse uploaded file → (SourceFile metadata, list of raw dict rows).

    Raises ParseError with §11 error codes on failure.
    """
    # 1. Size cap (before doing any work)
    size = len(file_bytes)
    if size == 0:
        raise ParseError("EMPTY_FILE", "上传的文件大小为 0")
    if size > MAX_FILE_BYTES:
        raise ParseError(
            "SIZE_EXCEEDED",
            f"文件 {size / 1024 / 1024:.1f} MB 超出 10 MB 上限",
        )

    # 2. Format detection
    fmt = detect_format(filename, file_bytes[:1024])

    # 3. Per-format parse
    encoding = "utf-8"
    actual_sheet = None
    if fmt == "xlsx":
        rows, actual_sheet = _parse_xlsx(file_bytes, sheet_name)
    else:
        text, encoding = detect_and_decode(file_bytes)
        if fmt == "json":
            rows = _parse_json(text)
        elif fmt == "jsonl":
            rows = _parse_jsonl(text)
        elif fmt == "txt":
            rows = _parse_txt(text)
        elif fmt == "csv":
            rows = _parse_csv(text)
        else:
            raise ParseError("PARSE_FORMAT_UNKNOWN", f"未支持的格式 {fmt}")

    # 4. Row-count cap
    if not rows:
        raise ParseError("EMPTY_FILE", "解析后无数据")
    if len(rows) > MAX_ROWS:
        raise ParseError(
            "ROW_EXCEEDED",
            f"行数 {len(rows)} 超出 {MAX_ROWS} 上限,请先离线抽样",
        )

    # 5. Build SourceFile metadata
    sf = SourceFile(
        filename=_safe_filename(filename),
        format=fmt,
        size_bytes=size,
        row_count=len(rows),
        encoding=encoding,
        sheet_name=actual_sheet,
        sample=[_truncate_row(r) for r in rows[:5]],
        uploaded_at=datetime.now(),
    )
    return sf, rows


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _safe_filename(name: str) -> str:
    """Strip path components and any leading dots."""
    base = name.replace("\\", "/").split("/")[-1]
    base = base.lstrip(".") or "unnamed"
    return base[:255]


def _truncate_row(row: dict, max_field_len: int = 200) -> dict:
    """Truncate any string fields for the sample preview (UI uses this)."""
    return {
        k: (v[:max_field_len] + "..." if isinstance(v, str) and len(v) > max_field_len else v)
        for k, v in row.items()
    }
